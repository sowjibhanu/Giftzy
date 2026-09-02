#!/usr/bin/env python3
"""Convert the GiftZy Investment workbook into install/seed.sql.

The workbook keeps several independent tables per worksheet (partner-funded
expenditure blocks, sale-account blocks, cash blocks, side blocks in far
columns), so every block is declared explicitly below with its row range and
column meaning.
"""
import datetime
import sys

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "GiftZy_Investment.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "seed.sql"

sales_rows = []
expense_rows = []
investment_rows = []
monthly_rows = []


def sql(v):
    if v is None or v == "":
        return "NULL"
    if isinstance(v, (int, float)):
        return repr(v)
    return "'" + str(v).replace("\\", "\\\\").replace("'", "''").strip() + "'"


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    try:
        return round(float(str(v).replace(",", "").strip()), 2)
    except ValueError:
        return None


def date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None


def text(v, limit=500):
    if v is None or v == "":
        return None
    return " ".join(str(v).split())[:limit]


def is_total(v):
    return text(v) is not None and text(v).lower().rstrip(":s") in ("total", "totals")


wb = openpyxl.load_workbook(XLSX, data_only=True)


# ---------------------------------------------------------------- sales sheets
def load_sales(sheet, first_row, customer_col, payment_col, pending_col):
    ws = wb[sheet]
    # The sheets write the date once per day and leave it blank on the rows that
    # follow, so an empty date means "same day as the row above". A Total row
    # closes a month block, after which the next date is written again.
    last_date = None
    for r in range(first_row, ws.max_row + 1):
        item = text(ws.cell(r, 1).value, 255)
        if not item or is_total(item) or item.lower() == "item":
            if item and is_total(item):
                last_date = None
            continue
        qty = num(ws.cell(r, 2).value) or 0
        sp = num(ws.cell(r, 3).value) or 0
        total = num(ws.cell(r, 4).value)
        cp = num(ws.cell(r, 5).value) or 0
        profit = num(ws.cell(r, 6).value)
        if total is None:
            total = round(qty * sp, 2)
        if profit is None:
            profit = round(total - qty * cp, 2)
        customer = text(ws.cell(r, customer_col).value, 255)
        pay_raw = (text(ws.cell(r, payment_col).value) or "").lower()
        if "online" in pay_raw or "phone" in pay_raw or "gpay" in pay_raw:
            pay = "Online"
        elif "cash" in pay_raw:
            pay = "Cash"
        elif "pend" in pay_raw:
            pay = "Pending"
        else:
            pay = "Other"
        pending = num(ws.cell(r, pending_col).value) or 0 if pending_col else 0
        notes = text(ws.cell(r, payment_col).value) if pay == "Other" else None
        d = date(ws.cell(r, 7).value)
        if d:
            last_date = d
        else:
            d = last_date
        sales_rows.append(
            (d, item, qty, sp, cp, total, profit,
             customer, pay, pending, notes, sheet)
        )


load_sales("2025 Sales", 2, 8, 9, None)
load_sales("Jan-Aug", 2, 8, 9, 10)
load_sales("SeptSale", 4, 8, 9, 10)


# ----------------------------------------------------------- expenditure sheet
EXP = wb["ExpendituresSettled"]


# Partner money per investment tranche, added up from the rows themselves so the
# summary entries and the expense rows can never drift apart.
tranche_totals = {}


def add_expense(d, purpose, amount, source, category=None, comments=None, settled=0,
                tranche=None):
    amount = num(amount)
    purpose = text(purpose, 255)
    if not purpose or amount is None or amount == 0 or is_total(purpose):
        return
    expense_rows.append((d, purpose, category, amount, source, settled,
                         comments, "ExpendituresSettled"))
    if tranche and source in ("Sowji", "Lavanya"):
        tranche_totals[tranche] = tranche_totals.get(tranche, 0) + amount


def partner_block(first, last, settled, category=None, category_col=None,
                  date_col=1, purpose_col=2, sowji_col=3, lavanya_col=4,
                  extra_col=None, extra_source="Shop", tranche=None):
    """Blocks that split one purpose across the two partners (self money)."""
    for r in range(first, last + 1):
        purpose = text(EXP.cell(r, purpose_col).value, 255)
        if not purpose or is_total(purpose) or "settle up" in purpose.lower():
            continue
        d = date(EXP.cell(r, date_col).value) if date_col else None
        comments = text(EXP.cell(r, 6).value) if date_col == 1 else None
        cat = category or (text(EXP.cell(r, category_col).value, 100)
                           if category_col else None)
        add_expense(d, purpose, EXP.cell(r, sowji_col).value, "Sowji", cat, comments, settled, tranche)
        add_expense(d, purpose, EXP.cell(r, lavanya_col).value, "Lavanya", cat, comments, settled, tranche)
        if extra_col:
            add_expense(d, purpose, EXP.cell(r, extra_col).value, extra_source, cat, comments, settled)


def account_block(first, last, skip_purposes=()):
    """Date | Purpose | Amount blocks - money spent out of the sale account."""
    for r in range(first, last + 1):
        purpose = text(EXP.cell(r, 2).value, 255)
        if not purpose or is_total(purpose) or purpose in skip_purposes:
            continue
        add_expense(date(EXP.cell(r, 1).value), purpose, EXP.cell(r, 3).value, "Account")


partner_block(2, 11, settled=1, category_col=7, tranche="Initial Expenditures")
partner_block(16, 107, settled=1, category_col=7, tranche="July to Oct")
partner_block(114, 150, settled=1, category_col=7, tranche="Oct to Dec")

# Cash box block: Amount | Purpose
for r in range(156, 164):
    add_expense(None, EXP.cell(r, 2).value, EXP.cell(r, 1).value, "Cash", "Cash box")

account_block(165, 249)
# Columns G:H of this block hold the investment summary, not a category.
partner_block(254, 304, settled=1, category="Shop build 2026", extra_col=5,
              tranche="2026 Investments")

partner_block(310, 323, settled=1, category="Mumbai trip", tranche="Mumbai Investment",
              date_col=None, purpose_col=1, sowji_col=2, lavanya_col=3)
partner_block(327, 347, settled=1, category="Stock purchase Mumbai", tranche="Mumbai Investment",
              date_col=None, purpose_col=1, sowji_col=2, lavanya_col=3)
partner_block(351, 363, settled=1, category="Shop setup / misc", tranche="Mumbai Investment",
              date_col=None, purpose_col=1, sowji_col=2, lavanya_col=3)

# Row 464 column C carries the Jan-Mar block total, not a new expense.
account_block(381, 454)
account_block(465, 578)
partner_block(464, 578, settled=0, category="Self expenditure Jul-Aug 2026",
              tranche="July to Aug 2026",
              date_col=None, purpose_col=5, sowji_col=6, lavanya_col=7)


# ---------------------------------------------------------------- amounts sheet
AM = wb["Amounts"]

# Investment summary block (G254:H259 on the expenditure sheet is the same data).
# The amounts come from the partner rows of the table each tranche covers, not from
# the sheet's Total cells, which are rounded and mix in money Giftzy paid itself.
for r in range(2, 7):
    purpose = text(AM.cell(r, 6).value, 255)
    if not purpose or is_total(purpose) or purpose not in tranche_totals:
        continue
    investment_rows.append((None, purpose, tranche_totals[purpose], "Other",
                            "Partner money in the " + purpose + " table"))

# The July-August self expenditure table (E463:G475) has no line in the Amounts
# summary block, so it goes in as one entry of its own.
if tranche_totals.get("July to Aug 2026"):
    investment_rows.append((None, "July to Aug 2026", tranche_totals["July to Aug 2026"],
                            "Other", "Partner money in the July-August self expenditure table"))

# Cash spends listed in the far-right column of the Amounts sheet
for r in range(3, 21):
    purpose = text(AM.cell(r, 11).value, 255)
    amount = num(AM.cell(r, 12).value)
    if purpose and amount and not is_total(purpose):
        expense_rows.append((None, purpose, "Cash box", amount, "Cash", 0,
                             "From the Amounts sheet cash list", "Amounts"))

MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july", "august",
     "september", "october", "november", "december"], start=1)}


def month_block(first, last, year, name_col, online_col, cash_col):
    for r in range(first, last + 1):
        raw = (text(AM.cell(r, name_col).value) or "").replace("Sale", "").strip().lower()
        if raw[:3] == "feb":
            raw = "february"
        if raw not in MONTHS:
            continue
        monthly_rows.append((
            "%d-%02d" % (year, MONTHS[raw]),
            num(AM.cell(r, online_col).value) or 0,
            num(AM.cell(r, cash_col).value) or 0,
        ))


month_block(2, 6, 2025, 1, 2, 3)
month_block(11, 19, 2026, 5, 6, 7)


# ------------------------------------------------------------------- write file
with open(OUT, "w") as f:
    f.write("-- Generated from the GiftZy Investment workbook by install/make_seed.py\n")
    f.write("SET NAMES utf8mb4;\n\n")
    f.write("-- Re-importing replaces the workbook data.\n")
    for t in ("sales", "expenses", "investments", "monthly_collections"):
        f.write("DELETE FROM %s;\n" % t)
    f.write("\n")

    f.write("INSERT INTO sales (sale_date,item,qty,selling_price,cost_price,total_amount,"
            "profit,customer,payment_type,pending_amount,notes,source_sheet) VALUES\n")
    f.write(",\n".join("(" + ",".join(sql(v) for v in row) + ")" for row in sales_rows))
    f.write(";\n\n")

    f.write("INSERT INTO expenses (expense_date,purpose,category,amount,fund_source,"
            "settled,comments,source_sheet) VALUES\n")
    f.write(",\n".join("(" + ",".join(sql(v) for v in row) + ")" for row in expense_rows))
    f.write(";\n\n")

    f.write("INSERT INTO investments (inv_date,purpose,amount,fund_source,notes) VALUES\n")
    f.write(",\n".join("(" + ",".join(sql(v) for v in row) + ")" for row in investment_rows))
    f.write(";\n\n")

    # Sale entries are what move the balances now, so each month's stored figure
    # becomes the difference between the sheet's collected amount and what the
    # sale entries of that month add up to.
    from_sales = {}
    for r in sales_rows:
        if not r[0]:
            continue  # undated sales are left out of the balances
        period = r[0][:7]
        collected = round(r[5] - (r[9] or 0), 2)
        channel = 1 if r[8] != "Cash" else 2
        month = from_sales.setdefault(period, [0.0, 0.0])
        month[channel - 1] += collected

    adjust_rows = []
    for period, online, cash in monthly_rows:
        sales_online, sales_cash = from_sales.get(period, (0.0, 0.0))
        adjust_rows.append((period,
                            round(online - sales_online, 2),
                            round(cash - sales_cash, 2)))
    f.write("INSERT INTO monthly_collections (period,online_amount,cash_amount) VALUES\n")
    f.write(",\n".join("(" + ",".join(sql(v) for v in row) + ")" for row in adjust_rows))
    f.write(";\n\n")

    # Openings that make the running balances land on the workbook's Amounts
    # sheet figures (account 38,833.69 and cash 10,720).
    acct_in = sum(r[1] for r in adjust_rows) + sum(m[0] for m in from_sales.values())
    cash_in = sum(r[2] for r in adjust_rows) + sum(m[1] for m in from_sales.values())
    acct_out = sum(r[3] for r in expense_rows if r[4] == "Account")
    cash_out = sum(r[3] for r in expense_rows if r[4] == "Cash")
    f.write("REPLACE INTO settings (name, value) VALUES\n")
    f.write("  ('opening_account', %s),\n" % round(38833.69 - acct_in + acct_out, 2))
    f.write("  ('opening_cash', %s);\n" % round(10720 - cash_in + cash_out, 2))

print("sales", len(sales_rows))
print("expenses", len(expense_rows), "total", round(sum(r[3] for r in expense_rows), 2))
print("  account", round(sum(r[3] for r in expense_rows if r[4] == "Account"), 2))
print("  sowji", round(sum(r[3] for r in expense_rows if r[4] == "Sowji"), 2))
print("  lavanya", round(sum(r[3] for r in expense_rows if r[4] == "Lavanya"), 2))
print("  cash", round(sum(r[3] for r in expense_rows if r[4] == "Cash"), 2))
print("investments", len(investment_rows), "total", sum(r[2] for r in investment_rows))
print("monthly", len(monthly_rows))
print("sales total", round(sum(r[5] for r in sales_rows), 2),
      "profit", round(sum(r[6] for r in sales_rows), 2))
